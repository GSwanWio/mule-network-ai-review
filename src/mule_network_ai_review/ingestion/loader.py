from collections import Counter
from io import BytesIO
from pathlib import Path
from typing import BinaryIO
from zipfile import BadZipFile, ZipFile

import pandas as pd
from openpyxl import load_workbook

from mule_network_ai_review.ingestion.models import (
	ValidationSummary,
	WorkbookPackage,
	WorkbookValidationError,
)
from mule_network_ai_review.ingestion.schema import load_workbook_contract

MAX_WORKBOOK_BYTES = 50 * 1024 * 1024
MAX_UNCOMPRESSED_BYTES = 250 * 1024 * 1024
PROHIBITED_ARCHIVE_PATHS = (
	"xl/vbaproject.bin",
	"xl/embeddings/",
	"xl/externallinks/",
)


def _read_source_bytes(source: bytes | bytearray | str | Path | BinaryIO) -> bytes:
	if isinstance(source, bytes):
		payload = source
	elif isinstance(source, bytearray):
		payload = bytes(source)
	elif isinstance(source, (str, Path)):
		payload = Path(source).read_bytes()
	elif hasattr(source, "read"):
		if hasattr(source, "seek"):
			source.seek(0)
		payload = source.read()
		if hasattr(source, "seek"):
			source.seek(0)
	else:
		raise TypeError("Workbook source must be bytes, a path, or a binary stream.")

	if not payload:
		raise WorkbookValidationError(["The workbook is empty."])
	if len(payload) > MAX_WORKBOOK_BYTES:
		raise WorkbookValidationError(["The workbook exceeds the permitted upload size."])
	return payload


def _validate_xlsx_container(payload: bytes) -> None:
	try:
		with ZipFile(BytesIO(payload)) as archive:
			archive_paths = [entry.filename.lower() for entry in archive.infolist()]
			uncompressed_bytes = sum(entry.file_size for entry in archive.infolist())
	except BadZipFile as error:
		raise WorkbookValidationError(["The upload is not a valid XLSX container."]) from error

	if uncompressed_bytes > MAX_UNCOMPRESSED_BYTES:
		raise WorkbookValidationError(["The workbook expands beyond the permitted size."])
	if any(
		path == prohibited_path or path.startswith(prohibited_path)
		for path in archive_paths
		for prohibited_path in PROHIBITED_ARCHIVE_PATHS
	):
		raise WorkbookValidationError(
			["The workbook contains prohibited executable, embedded, or external content."]
		)


def _validate_formula_free(payload: bytes) -> None:
	workbook = load_workbook(
		BytesIO(payload),
		read_only=True,
		data_only=False,
		keep_links=False,
	)
	formula_count = sum(
		1
		for worksheet in workbook.worksheets
		for row in worksheet.iter_rows()
		for cell in row
		if cell.data_type == "f"
	)
	workbook.close()
	if formula_count:
		raise WorkbookValidationError(
			["The workbook contains formulas; only materialized values are permitted."]
		)


def _blank_mask(series: pd.Series) -> pd.Series:
	return series.isna() | series.fillna("").astype(str).str.strip().eq("")


def _flag(value: object) -> int | None:
	if pd.isna(value):
		return None
	if isinstance(value, bool):
		return int(value)
	try:
		return int(value)
	except (TypeError, ValueError):
		return None


def _validate_sheet_contracts(
	frames: dict[str, pd.DataFrame],
	contract: dict,
) -> list[str]:
	issues: list[str] = []
	for sheet_name, sheet_contract in contract["sheets"].items():
		frame = frames[sheet_name]
		expected_columns = sheet_contract["columns"]
		actual_columns = frame.columns.tolist()
		if actual_columns != expected_columns:
			missing_columns = sorted(set(expected_columns) - set(actual_columns))
			unexpected_columns = sorted(set(actual_columns) - set(expected_columns))
			if missing_columns:
				issues.append(
					f"Sheet {sheet_name} is missing {len(missing_columns)} required columns."
				)
			if unexpected_columns:
				issues.append(
					f"Sheet {sheet_name} contains {len(unexpected_columns)} unexpected columns."
				)
			if not missing_columns and not unexpected_columns:
				issues.append(f"Sheet {sheet_name} has an invalid column order.")
			continue

		key_columns = sheet_contract["key_columns"]
		blank_key_rows = frame[key_columns].apply(_blank_mask).any(axis=1).sum()
		duplicate_key_rows = frame.duplicated(subset=key_columns, keep=False).sum()
		if blank_key_rows:
			issues.append(f"Sheet {sheet_name} contains {blank_key_rows} blank key rows.")
		if duplicate_key_rows:
			issues.append(f"Sheet {sheet_name} contains {duplicate_key_rows} duplicate key rows.")
	return issues


def _validate_manifest(frames: dict[str, pd.DataFrame], contract: dict) -> list[str]:
	issues: list[str] = []
	manifest = frames["manifest"]
	if len(manifest) != 1:
		return ["The manifest must contain exactly one data row."]

	manifest_row = manifest.iloc[0]
	protection = contract["protection"]
	if str(manifest_row["data_classification"]).strip() != protection["data_classification"]:
		issues.append("The manifest data classification is not approved.")
	if _flag(manifest_row["contains_original_identifiers"]) != protection[
		"contains_original_identifiers"
	]:
		issues.append("The manifest indicates that original identifiers may be present.")
	if _flag(manifest_row["reidentification_mapping_included"]) != protection[
		"reidentification_mapping_included"
	]:
		issues.append("The manifest indicates that a re-identification mapping may be present.")

	row_count_columns = {
		"network_summary": "network_rows",
		"nodes": "node_rows",
		"relationships": "relationship_rows",
		"customer_metrics": "customer_metric_rows",
		"counterparty_local": "local_counterparty_metric_rows",
		"counterparty_intl": "international_counterparty_metric_rows",
	}
	for sheet_name, count_column in row_count_columns.items():
		if _flag(manifest_row[count_column]) != len(frames[sheet_name]):
			issues.append(f"Manifest row count does not reconcile for sheet {sheet_name}.")
	return issues


def _network_node_keys(frame: pd.DataFrame) -> set[tuple[str, str]]:
	return set(
		zip(
			frame["network_id"].astype(str),
			frame["node_id"].astype(str),
			strict=True,
		)
	)


def _network_subject_keys(
	frame: pd.DataFrame,
	subject_column: str,
) -> set[tuple[str, str]]:
	return set(
		zip(
			frame["network_id"].astype(str),
			frame[subject_column].astype(str),
			strict=True,
		)
	)


def _validate_graph(frames: dict[str, pd.DataFrame], contract: dict) -> list[str]:
	issues: list[str] = []
	networks = frames["network_summary"]
	nodes = frames["nodes"]
	relationships = frames["relationships"]
	network_ids = set(networks["network_id"].astype(str))

	if set(nodes["network_id"].astype(str)) != network_ids:
		issues.append("Node network coverage does not match the network summary.")
	if set(relationships["network_id"].astype(str)) != network_ids:
		issues.append("Relationship network coverage does not match the network summary.")

	unexpected_node_types = set(nodes["node_type"].astype(str)) - set(
		contract["allowed_node_types"]
	)
	if unexpected_node_types:
		issues.append(f"Nodes contain {len(unexpected_node_types)} unsupported node types.")

	unexpected_relationship_types = set(relationships["relationship_type"].astype(str)) - set(
		contract["allowed_relationship_types"]
	)
	if unexpected_relationship_types:
		issues.append(
			f"Relationships contain {len(unexpected_relationship_types)} unsupported types."
		)

	token_requirements = {
		"CUSTOMER": ["customer_token"],
		"EID": ["eid_token"],
		"COUNTERPARTY": [
			"counterparty_token",
			"counterparty_key_token",
			"account_token",
		],
	}
	for node_type, token_columns in token_requirements.items():
		typed_nodes = nodes.loc[nodes["node_type"].astype(str) == node_type]
		for token_column in token_columns:
			blank_count = _blank_mask(typed_nodes[token_column]).sum()
			if blank_count:
				issues.append(
					f"Node type {node_type} contains {blank_count} blank {token_column} values."
				)

	seed_nodes = nodes.loc[nodes["is_seed_customer"].map(_flag) == 1]
	if len(seed_nodes) != len(networks):
		issues.append("The workbook must contain exactly one seed node per network.")

	node_keys = _network_node_keys(nodes)
	source_keys = set(
		zip(
			relationships["network_id"].astype(str),
			relationships["source_node_id"].astype(str),
			strict=True,
		)
	)
	target_keys = set(
		zip(
			relationships["network_id"].astype(str),
			relationships["target_node_id"].astype(str),
			strict=True,
		)
	)
	if source_keys - node_keys:
		issues.append("Relationships contain source nodes that are not present in the node sheet.")
	if target_keys - node_keys:
		issues.append("Relationships contain target nodes that are not present in the node sheet.")
	return issues


def _validate_metric_coverage(frames: dict[str, pd.DataFrame]) -> list[str]:
	issues: list[str] = []
	nodes = frames["nodes"]
	customer_nodes = nodes.loc[nodes["node_type"].astype(str) == "CUSTOMER"]
	counterparty_nodes = nodes.loc[nodes["node_type"].astype(str) == "COUNTERPARTY"]

	expected_customer_keys = _network_subject_keys(customer_nodes, "customer_token")
	actual_customer_keys = _network_subject_keys(frames["customer_metrics"], "customer_token")
	if expected_customer_keys != actual_customer_keys:
		issues.append("Customer metric coverage does not match discovered customer nodes.")

	expected_counterparty_keys = _network_subject_keys(
		counterparty_nodes,
		"counterparty_token",
	)
	for sheet_name in ("counterparty_local", "counterparty_intl"):
		actual_counterparty_keys = _network_subject_keys(
			frames[sheet_name],
			"counterparty_token",
		)
		if expected_counterparty_keys != actual_counterparty_keys:
			issues.append(
				f"Counterparty metric coverage does not match nodes for sheet {sheet_name}."
			)
	return issues


def _shared_subject_count(frame: pd.DataFrame, token_column: str) -> int:
	grouped = frame.groupby(token_column, dropna=False)["network_id"].nunique()
	return int((grouped > 1).sum())


def _build_validation_summary(
	frames: dict[str, pd.DataFrame],
	contract: dict,
) -> ValidationSummary:
	manifest_row = frames["manifest"].iloc[0]
	nodes = frames["nodes"]
	node_counts = Counter(nodes["node_type"].astype(str))
	return ValidationSummary(
		schema_version=contract["schema_version"],
		export_run_id=str(manifest_row["export_run_id"]),
		network_count=len(frames["network_summary"]),
		customer_node_count=node_counts.get("CUSTOMER", 0),
		eid_node_count=node_counts.get("EID", 0),
		counterparty_node_count=node_counts.get("COUNTERPARTY", 0),
		relationship_count=len(frames["relationships"]),
		shared_customer_count=_shared_subject_count(
			nodes.loc[nodes["node_type"].astype(str) == "CUSTOMER"],
			"customer_token",
		),
		shared_counterparty_count=_shared_subject_count(
			nodes.loc[nodes["node_type"].astype(str) == "COUNTERPARTY"],
			"counterparty_token",
		),
		sheet_row_counts={
			sheet_name: len(frame)
			for sheet_name, frame in frames.items()
		},
	)


def load_workbook_package(
	source: bytes | bytearray | str | Path | BinaryIO,
) -> WorkbookPackage:
	payload = _read_source_bytes(source)
	_validate_xlsx_container(payload)
	_validate_formula_free(payload)
	contract = load_workbook_contract()

	excel_file = pd.ExcelFile(BytesIO(payload), engine="openpyxl")
	expected_sheets = list(contract["sheets"])
	actual_sheets = excel_file.sheet_names
	missing_sheets = sorted(set(expected_sheets) - set(actual_sheets))
	unexpected_sheets = sorted(set(actual_sheets) - set(expected_sheets))
	issues: list[str] = []
	if missing_sheets:
		issues.append(f"The workbook is missing {len(missing_sheets)} required sheets.")
	if unexpected_sheets:
		issues.append(f"The workbook contains {len(unexpected_sheets)} unexpected sheets.")
	if issues:
		raise WorkbookValidationError(issues)

	frames = {
		sheet_name: pd.read_excel(
			excel_file,
			sheet_name=sheet_name,
			dtype=object,
		).rename(columns=lambda column: str(column).strip())
		for sheet_name in expected_sheets
	}

	issues.extend(_validate_sheet_contracts(frames, contract))
	if issues:
		raise WorkbookValidationError(issues)

	issues.extend(_validate_manifest(frames, contract))
	issues.extend(_validate_graph(frames, contract))
	issues.extend(_validate_metric_coverage(frames))
	if issues:
		raise WorkbookValidationError(issues)

	return WorkbookPackage(
		frames=frames,
		validation_summary=_build_validation_summary(frames, contract),
	)
